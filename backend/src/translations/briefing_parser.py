# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Parsing of the customer's translation spreadsheet (Copy Sheet + TM)."""

import io
import re

import openpyxl

from src.translations.markets import MARKETS, SOURCE_MARKET

_METADATA_LABELS = {
    "email": "email",
    "requestor": "requestor",
    "date email": "date_email",
    "due": "due",
    "notes": "notes",
}
_REQUEST_RE = re.compile(r"^request\s*nr", re.IGNORECASE)
_MAX_RE = re.compile(r"max\s*(\d+)", re.IGNORECASE)


def _cell(value) -> str:
    return "" if value is None else str(value).strip()


def _char_limit(label: str) -> int | None:
    m = _MAX_RE.search(label)
    return int(m.group(1)) if m else None


def _field_name(label: str) -> str:
    """Strips the '(MAX n ...)' hint to get a clean field name."""
    return re.sub(r"\s*\(.*?\)\s*$", "", label).strip()


def _load_sheet_rows(file_bytes: bytes, sheet_name: str) -> list[list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def list_sheets(file_bytes: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _market_columns(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Finds the header row and maps each market code to its column index.

    Returns (header_row_index, {market_code: col_index}).
    """
    for i, row in enumerate(rows[:8]):
        upper = [c.upper() for c in row]
        if SOURCE_MARKET in upper:
            mapping = {}
            for col, val in enumerate(upper):
                # Normalise NOR -> NO to match our market codes.
                code = "NO" if val == "NOR" else val
                if code in MARKETS:
                    mapping[code] = col
            if SOURCE_MARKET in mapping:
                return i, mapping
    raise ValueError("Could not find a market header row (expected an 'EN' column).")


def find_requests(file_bytes: bytes, sheet_name: str) -> list[dict]:
    """Returns the 'Request nr. X' markers found in a sheet."""
    rows = _load_sheet_rows(file_bytes, sheet_name)
    requests = []
    idx = 0
    for r, row in enumerate(rows):
        a = row[0] if row else ""
        if _REQUEST_RE.match(a):
            requests.append({"index": idx, "label": a, "row": r})
            idx += 1
    return requests


def parse_request(file_bytes: bytes, sheet_name: str, request_index: int) -> dict:
    """Parses one request into metadata + source (EN) segments."""
    rows = _load_sheet_rows(file_bytes, sheet_name)
    _, market_cols = _market_columns(rows)
    en_col = market_cols[SOURCE_MARKET]

    markers = [
        r for r, row in enumerate(rows) if row and _REQUEST_RE.match(row[0])
    ]
    if request_index >= len(markers):
        raise ValueError(f"Request index {request_index} not found.")
    start = markers[request_index]
    end = markers[request_index + 1] if request_index + 1 < len(markers) else len(rows)

    meta: dict[str, str] = {"request_label": rows[start][0]}
    segments: list[dict] = []
    current_block: str | None = None

    for row in rows[start + 1 : end]:
        a = row[0] if len(row) > 0 else ""
        b = row[1] if len(row) > 1 else ""
        en_val = row[en_col] if len(row) > en_col else ""

        a_key = a.lower()
        if a_key in _METADATA_LABELS:
            # Metadata value sits in the EN/source column (or col B as fallback).
            meta[_METADATA_LABELS[a_key]] = en_val or b
            continue

        # A non-empty col A that isn't metadata/request marks a new block.
        if a and not _REQUEST_RE.match(a):
            current_block = a

        # Field label lives in col B; skip rows without one.
        if not b:
            continue

        segments.append(
            {
                "block": current_block,
                "field": _field_name(b),
                "label": b,
                "char_limit": _char_limit(b),
                "text": en_val,
            }
        )

    name = meta.get("email") or meta.get("request_label") or f"{sheet_name} request"
    return {"name": name, "meta": meta, "segments": segments}


def parse_translation_memories(
    file_bytes: bytes, sheet_name: str = "Translation Memories"
) -> list[dict]:
    """Parses the TM sheet into glossary rows: {market, source, target}."""
    rows = _load_sheet_rows(file_bytes, sheet_name)
    header_idx, market_cols = _market_columns(rows)
    en_col = market_cols[SOURCE_MARKET]
    target_cols = {m: c for m, c in market_cols.items() if m != SOURCE_MARKET}

    entries: list[dict] = []
    for row in rows[header_idx + 1 :]:
        source = row[en_col] if len(row) > en_col else ""
        if not source:
            continue
        for market, col in target_cols.items():
            target = row[col] if len(row) > col else ""
            if target and target != source:
                entries.append(
                    {"market": market, "source": source, "target": target}
                )
    return entries
