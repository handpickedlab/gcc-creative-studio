# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Exports a briefing + its translations back to an xlsx (Copy Sheet layout)."""

import io

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.translations.markets import MARKETS, SOURCE_MARKET

_HEADER_FILL = PatternFill("solid", fgColor="1F2430")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LABEL_FONT = Font(bold=True)


def build_briefing_xlsx(briefing: dict, translations: list[dict]) -> bytes:
    """briefing = {name, meta, segments}; translations = [{market, segments}].

    Segments in each translation align by index with briefing['segments'].
    """
    segments = briefing.get("segments", [])
    meta = briefing.get("meta", {}) or {}

    # Column order: source first, then translated markets in MARKETS order.
    translated_markets = [t["market"] for t in translations]
    ordered = [m for m in MARKETS if m == SOURCE_MARKET or m in translated_markets]
    tr_by_market = {t["market"]: t.get("segments", []) for t in translations}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Header: col A (block) | col B (field) | market columns
    header = ["", ""] + ordered
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Metadata block
    ws.append([briefing.get("name", "")])
    ws.cell(row=ws.max_row, column=1).font = _LABEL_FONT
    for key, label in [
        ("email", "Email"),
        ("requestor", "Requestor"),
        ("date_email", "Date email"),
        ("due", "Due"),
        ("notes", "Notes"),
    ]:
        val = meta.get(key) or ""
        if val:
            ws.append([label, "", val])

    ws.append([])  # spacer

    # Segment rows
    prev_block = object()
    for i, seg in enumerate(segments):
        block = seg.get("block")
        col_a = block if block != prev_block else ""
        prev_block = block
        row = [col_a or "", seg.get("label", seg.get("field", ""))]
        for m in ordered:
            if m == SOURCE_MARKET:
                row.append(seg.get("text", ""))
            else:
                tr_segs = tr_by_market.get(m, [])
                row.append(tr_segs[i].get("text", "") if i < len(tr_segs) else "")
        ws.append(row)

    # Reasonable column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    for idx in range(len(ordered)):
        ws.column_dimensions[chr(ord("C") + idx)].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
